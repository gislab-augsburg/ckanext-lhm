#!/usr/bin/env python

import urllib.request as urllib
import urllib.parse
import json
import pprint
import sys
import csv
import pandas as pd
import codecs
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.reader.excel import warnings
import os
import requests
import time
import shutil
import __main__
# MB_CHANGE FOR DOWNLOAD BUTTON ###
import ckanext.lhm.tp as tp
###############################

# Suitable for template_v1.2.2.xlsx

# Defining Config File
# MB_CHANGE FOR DOWNLOAD BUTTON ###
def config():
    #if len(sys.argv) != 2:
    #    print("Please specify config file as argument, e.g.: python3 get-data.py ../config-demo.json")
    #    sys.exit()
    #else:
    #    arg_config = sys.argv[1]
    arg_config = '/srv/app/md_download/config-download.json'
#################################

    #Sensitive information are stored in a config file.
    #api_key is stored in the config and is called with the following:
    with open(arg_config, 'r') as config_file:
        config = json.load(config_file)
    api_key = config['api_key_harvester']
    base_url = config['base_url']

    return api_key, base_url



# Define request, get response as dict
def ckan_request(api_string):
    # Get config parameters
    config_ = config()
    api_key = config_[0]
    base_url = config_[1]
    # Define the request
    url = base_url + api_string
    headers = {'Authorization': api_key}
    print('URL:')
    print(url)
    print('HEADERS:')
    print(headers)
    response = requests.get(url, headers=headers)
    print('RESPONSE:')
    print(response)
    for ir in range(0, 10):
        if response.status_code == 200:
            response_dict = response.json()
            assert response_dict['success'] is True
            result = response_dict['result']
            return result
        else:
            print(f'* Request failed, status code {str(response.status_code)}')
            print(f'* {response.text}')
            print(f'* Trying again: {url}')
            time.sleep(2)
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                response_dict = response.json() 
                assert response_dict['success'] is True
                result = response_dict['result']
                print('* Success')
                return result

# Add BOM
def add_utf8_bom(filename):
    with codecs.open(filename, 'r', 'utf-8') as f:
        content = f.read()
    with codecs.open(filename, 'w', 'utf-8') as f2:
        f2.write('\ufeff')
        f2.write(content)
    return

# Add dropdown validation
def add_validation_dropdown(ws, coords, formula):
    val_fields = []
    # Use existing validation
    for dv in ws.data_validations.dataValidation:   
        val_fields.append(dv.sqref)       
        if dv.sqref == coords:
            dv.type = 'list'
            dv.formula1 = formula
            dv.allow_blank = False
            dv.error ='Bitte wählen Sie einen Wert aus der Drop-Down Liste'
            dv.errorTitle = 'Ungültiger Eintrag'
            dv.showErrorMessage = True
            dv_final = dv
    # Add new validation if no existing
    if coords not in val_fields:       
        cell = ws[coords]
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        dv.error ='Bitte wählen Sie einen Wert aus der Drop-Down Liste'
        dv.errorTitle = 'Ungültiger Eintrag'
        dv.showErrorMessage = True
        dv.add(cell)
        dv_final = dv
        
    return dv_final

# Add field-not-empty validation
def add_validation_mandatory(ws, coords):
    val_fields = []
    # Use existing validation
    for dv in ws.data_validations.dataValidation:   
        val_fields.append(dv.sqref)       
        if dv.sqref == coords:
            dv.type = 'textLength'
            dv.operator = 'greaterThan'
            dv.formula1 = '0'
            dv.allow_blank = False
            dv.error ='Dieses Feld darf nicht leer sein, ein Wert ist erforderlich'
            dv.errorTitle = 'Ungültiger Eintrag'
            dv.showErrorMessage = True
            dv_final = dv
    # Add new validation if no existing
    if coords not in val_fields:       
        cell = ws[coords]
        dv = DataValidation(type="textLength", operator="greaterThan", formula1='0', allow_blank=True)
        dv.error ='Dieses Feld darf nicht leer sein, ein Wert ist erforderlich'
        dv.errorTitle = 'Ungültiger Eintrag'
        dv.showErrorMessage = True
        dv.add(cell)
        dv_final = dv
        
    return dv_final
  

# Get GDP packages and their resources from package list via API and save them
# a) in json dumps
# b) in excel and csv files  
# All subfields, dicts and lists required for editing are resolved.

def packages_to_files(packages, limit, wdir, excel_template):

    # packages: list of package names or string of single package name
    # limit: maximum number of packages to process
    # wdir: work and output directory
    # excel_template: excel template to use

    # Look for output dir, create it if not existent
    if os.path.exists(wdir) == False:
        os.mkdir(wdir)
    if os.path.exists(f'{wdir}/json') == False:
        os.mkdir(f'{wdir}/json')
    if os.path.exists(f'{wdir}/excel') == False:
        os.mkdir(f'{wdir}/excel')

    # Get config parameters
    config_ = config()
    api_key = config_[0]
    base_url = config_[1]

    ## Check if packages is list or single package name
    if type(packages) == list:
        package_list = packages
    elif type(packages) == str:
        package_list = []
        package_list.append(packages)

    ## Loop through package list
    i = 0
    for id in package_list:
        if i < limit:
            ##
            #id = "historisches_luftbildmosaik_1984"
            #print('.')
            ##
            package_dict = ckan_request(f'/api/3/action/package_show?id={id}')
            df = pd.DataFrame([package_dict])
            type_ = df['type'][0]
            # MB_CHANGE FOR DOWNLOAD BUTTON ### add mobidam
            if type_ == 'geodatenpool' or type_ == 'mobidam':
            ###################################
                p_name = package_dict['name']
                # MB_CHANGE FOR DOWNLOAD BUTTON ### comment out the following
                #if __name__ == "__main__":
                #    print(p_name)
                #if "harvest_init.py" in __main__.__file__:
                #    print(p_name)
                # Open excel template, define sheets
                ###################################
                warnings.simplefilter(action='ignore')
                wb = load_workbook(filename=excel_template)
                dv_exists = False
                kw_exists = False
                ws = wb["GDP Metadaten"]
                ws_1 = wb["Datenverzeichnis"]
                ws_2 = wb["Katalogwerte"]
                ws_3 = wb["Dienste und Dokumente"]
                #ws = wb.active
                with open(f'{wdir}/json/{p_name}_PRE.json', 'w') as fp:
                    json.dump(package_dict, fp)

                # Delete other resource files (the current will be recreated, no more existent will not be created)
                if os.path.exists(f'{wdir}/excel/{p_name}') == True:
                    shutil.rmtree(f'{wdir}/excel/{p_name}')
                
                ### Resolve required subfields, dicts and lists and add to package_dict
                i = i + 1
                for key in df.keys():
                    value = df[key][0]
                    # Genauigkeit
                    if key == "genauigkeit":
                        genauigkeit_dict = json.loads(value, strict=False)
                        if 'lagegenauigkeit' in genauigkeit_dict.keys():
                            package_dict['genauigkeit.lagegenauigkeit'] = genauigkeit_dict['lagegenauigkeit']
                        else:
                            package_dict['genauigkeit.lagegenauigkeit'] = ''
                        if 'bemerkung' in genauigkeit_dict.keys():
                            package_dict['genauigkeit.bemerkung'] = genauigkeit_dict['bemerkung']
                        else:
                            package_dict['genauigkeit.bemerkung'] = ''
                    # Intranet
                    #if key == "intranet":
                    #    intranet_dict = json.loads(value, strict=False)
                    #    package_dict['intranet.fachverfahren'] = intranet_dict['fachverfahren']
                    #    package_dict['intranet.geoinfoweb'] = intranet_dict['geoinfoweb']
                    # Organization
                    if key == "organization":
                        package_dict['organization.title'] = value['title']
                    # Stadtarchiv
                    if key == "stadtarchiv":
                        stadtarchiv_dict = json.loads(value, strict=False)
                        package_dict['stadtarchiv.archivwuerdigkeit'] = stadtarchiv_dict['archivwuerdigkeit']
                        if 'begruendung' in stadtarchiv_dict.keys():
                            package_dict['stadtarchiv.begruendung'] = stadtarchiv_dict['begruendung']
                        else:
                            package_dict['stadtarchiv.begruendung'] = ''
                    # Groups - Main Categories
                    if key == "groups":
                        if len(value) > 0:
                            for j in range(len(value)):
                                if value[j]['title'] in tp.main_group_transpose:
                                    package_dict[f'groups.main.title'] = value[j]['title']
                    # Groups- Topics
                    k = -1
                    if key == "groups":
                        if len(value) > 0:
                            for j in range(len(value)):
                                if value[j]['title'] in tp.topic_group_transpose:
                                    k = k + 1
                                    package_dict[f'groups.topic.{k}.title'] = value[j]['title']
                    # LHM-extern - Nutzungsoptionen bei Einschränkungen
                    if key == "nutzungsoptionen_bei_einschraenkungen":
                        nutzungsoptionen_list = []
                        if value[0] == '{' and value[-1] == '}':
                            # more than 1 option
                            nu = value[1:][:-1]
                            number = len(nu.split(','))
                            for j in range(0,number):
                                nu2 = nu.split(',')[j]
                                nu3 = f'{nu2}'
                                nutzungsoptionen_list.append(nu3)
                        else:
                            # 1 option
                            nutzungsoptionen_list.append(value)
                        package_dict[f'nutzungsoptionen_list'] = nutzungsoptionen_list

                    # Nutzungshinweise
                    if key == "nutzungshinweise":
                        if len(value) > 0:
                            for j in range(len(value)):
                                package_dict[f'nutzungshinweise.{j}.stichwort'] = value[j]['stichwort']
                                package_dict[f'nutzungshinweise.{j}.hinweise'] = value[j]['hinweise']
                    # Resources
                    border_thin = Side(style='thin')
                    Border_thin = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
                    number_other_res = 0
                    m = 1
                    if key == "resources":
                        if len(value) > 0:
                            for j in range(len(value)):
                                package_dict[f'resources.{j}.url'] = value[j]['url']
                                package_dict[f'resources.{j}.name'] = value[j]['name']
                                if 'description' in value[j]:
                                    package_dict[f'resources.{j}.description'] = value[j]['description']
                                else:
                                    package_dict[f'resources.{j}.description'] = ''
                                package_dict[f'resources.{j}.format'] = value[j]['format']

                                # Check if resources available in datastore
                                #if value[j]['datastore_active'] == False:
                                    #print(f'resource {j}: datastore_active: false')
                                    #print('-------------------')

                                # Datenverzeichnis
                                if value[j]['format'] == 'oracle-db' and value[j]['position'] == 0:
                                    dv_exists = True
                                    res_id = value[j]['id']
                                    #print(f'/api/3/action/datastore_search?resource_id={res_id}')
                                    # for dv which has no table, datsore = false:
                                    if value[j]['datastore_active']:
                                        res_dict = ckan_request(f'/api/3/action/datastore_search?resource_id={res_id}')
                                        ## create json
                                        res_name = value[j]['name'].lower()
                                        with open(f'{wdir}/json/{id}.resources.{j}.dv.json', 'w') as fp:
                                            json.dump(res_dict, fp)
                                        ## create excel sheet
                                        #ws_1 = wb["Datenverzeichnis"]
                                        #m = 1
                                        for field in res_dict['fields']:
                                            if field['id'] != '_id':
                                                m = m + 1
                                                # Attribut
                                                Att = field['id']
                                                coords = f'A{m}'
                                                cell = ws_1[coords]
                                                cell.value = Att
                                                cell.border = Border_thin
                                                # Datentyp
                                                Dat = field['type']
                                                for key, val in tp.dv_datentyp_transpose.items():
                                                    if val == Dat:
                                                        Dat = key
                                                coords = f'B{m}'
                                                cell = ws_1[coords]
                                                cell.value = Dat
                                                cell.border = Border_thin
                                                dv = add_validation_dropdown(ws_1, coords, '=Wertelisten!$P$2:$P$8')
                                                ws_1.add_data_validation(dv)
                                                # Bedingung
                                                Bed = field['info']['condition']
                                                coords = f'C{m}'
                                                cell = ws_1[coords]
                                                cell.value = Bed
                                                cell.border = Border_thin
                                                # Beschreibung
                                                Bes = field['info']['notes']
                                                coords = f'D{m}'
                                                cell = ws_1[coords]
                                                cell.value = Bes
                                                cell.border = Border_thin
                                                # Bemerkung
                                                Bem = field['info']['label']
                                                coords = f'E{m}'
                                                cell = ws_1[coords]
                                                cell.value = Bem
                                                cell.border = Border_thin
                                                # Open Data
                                                Ope = field['info']['opendata']
                                                coords = f'F{m}'
                                                cell = ws_1[coords]
                                                cell.value = Ope
                                                cell.border = Border_thin
                                                dv = add_validation_dropdown(ws_1, coords, '=Wertelisten!$Q$2:$Q$4')
                                                ws_1.add_data_validation(dv)
                                            
                                
                                # Katalogwerte
                                #elif value[j]['format'] in ['xlsx', 'XLSX', 'csv', 'CSV'] and value[j]['position'] == 1 and value[j]['datastore_active'] == True:
                                #elif value[j]['format'] in ['xlsx', 'XLSX', 'csv', 'CSV'] and value[j]['position'] == 1:
                                # Selection ok?
                                #elif value[j]['format'] in ['xlsx', 'XLSX', 'csv', 'CSV'] and value[j]['position'] != 0:
                                elif value[j]['name'] == "Katalogwerte" and value[j]['format'] in ['xlsx', 'XLSX', 'csv', 'CSV']:
                                    kw_exists = True
                                    res_id = value[j]['id']
                                    # HTTP Error 404: NOT FOUND if immediatly, sleep needed (???):
                                    #if not __name__ == "__main__":
                                        #time.sleep(2)
                                    res_dict = ckan_request(f'/api/3/action/datastore_search?resource_id={res_id}')
                                    ## create json
                                    res_name = value[j]['name'].lower()
                                    with open(f'{wdir}/json/{id}.resources.{j}.kw.json', 'w') as fp:
                                        json.dump(res_dict, fp)
                                    ## create excel sheet
                                    #ws_2 = wb["Katalogwerte"]
                                    records = res_dict['records']
                                    m_ = 0
                                    for key in records[0].keys():
                                        col_ = chr(65 + m_)
                                        m_ = m_ + 1
                                        cell = ws_2[f'{col_}1']
                                        cell.value = key
                                        cell.fill = PatternFill("solid", start_color="8EA9DB")
                                        cell.border = Border_thin
                                        n = 1
                                    
                                        for nr in range(len(records)):
                                            n = n + 1
                                            value__ = records[nr][key]
                                            row__ = n
                                            cell = ws_2[f'{col_}{row__}']
                                            cell.value = value__
                                            cell.border = Border_thin

                                # Other Resources
                                else:
                                    number_other_res = number_other_res + 1

                                    format = value[j]['format']
                                    #if format.lower() in ['pdf', 'wcs', 'wfs', 'wms', 'wmts', 'json', 'xml/gml', 'ArcGIS Kartenservice']:
                                    #    format = format
                                    #else: format = 'other'
                                    ## Create excel sheet
                                    #ws_3 = wb["Dienste und Dokumente"]
                                    line__ =  str(number_other_res + 1)
                                    # Name
                                    coords = f'A{line__}'
                                    cell = ws_3[coords]
                                    cell.value = value[j]['name']
                                    cell.border = Border_thin
                                    dv = add_validation_mandatory(ws_3, coords)
                                    ws_3.add_data_validation(dv)
                                    # Beschreibung
                                    coords = f'B{line__}'
                                    cell = ws_3[coords]
                                    cell.value = value[j]['description']
                                    cell.border = Border_thin
                                    # Format
                                    coords = f'C{line__}'
                                    cell = ws_3[coords]
                                    cell.value = value[j]['format']
                                    cell.border = Border_thin
                                    dv = add_validation_dropdown(ws_3, coords, '=Wertelisten!$R$2:$R$11')
                                    ws_3.add_data_validation(dv)
                                    # URL
                                    coords = f'D{line__}'
                                    cell = ws_3[coords]
                                    cell.value = value[j]['url']
                                    cell.border = Border_thin
                                    # Download resource files if uploaded as file and now in mdk-pvc
                                    if value[j]['url_type'] == 'upload':
                                        url_res = str(value[j]['url'])
                                        file_name = url_res.split('/download/')[1]
                                        if os.path.exists(f'{wdir}/excel/{p_name}') == False:
                                            os.mkdir(f'{wdir}/excel/{p_name}')
                                        file_path = f'{wdir}/excel/{p_name}/{file_name}'
                                        #print(url_res)
                                        response_res = requests.get(url_res)
                                        with open(file_path, mode="wb") as file:
                                            file.write(response_res.content)

                                    
                    # Tags
                    if key == "tags":
                        if len(value) > 0:
                            for j in range(len(value)):
                                package_dict[f'tags.{j}.name'] = value[j]['name']
                    # Relationships
                    if key == "relationships_as_subject":
                        if len(value) > 0:
                            for j in range(len(value)):
                                package_dict[f'relationships_as_subject.{j}'] = value[j]
                                package_dict[f'relationships_as_object.{j}'] = value[j]


                # Add data validation to empty fields
                # DV-Datentyp
                for pos in range(m + 1, 201):
                    coords = f'B{str(pos)}'
                    cell = ws_1[coords]
                    dv = add_validation_dropdown(ws_1, coords, '=Wertelisten!$P$2:$P$8')
                    ws_1.add_data_validation(dv)

                # DV-OpenData
                for pos in range(m + 1, 201):
                    coords = f'F{str(pos)}'
                    cell = ws_1[coords]
                    dv = add_validation_dropdown(ws_1, coords, '=Wertelisten!$Q$2:$Q$4')
                    ws_1.add_data_validation(dv)

                # Dienste-Format
                for pos in range(number_other_res + 1, 201):
                    coords = f'C{str(pos)}'
                    cell = ws_3[coords]
                    dv = add_validation_dropdown(ws_3, coords, '=Wertelisten!$R$2:$R$11')
                    ws_3.add_data_validation(dv)

                # Delete Sheets if no Datenverzeichnis or Katalogwerte
                #if dv_exists == False:
                    #ws_1 = wb["Datenverzeichnis"]
                    #wb.remove(ws_1)
                    #print('  delete DV')
                #if kw_exists == False:
                    #ws_2 = wb["Katalogwerte"]
                    #wb.remove(ws_2)
                    #print('  delete KW')
                dv_exists = False
                kw_exists = False
                #print(wb.sheetnames)
                    
                # Create json 
                pack_name = package_dict['name']
                with open(f'{wdir}/json/{pack_name}.json', 'w') as fp:
                    json.dump(package_dict, fp)
                #print(package_dict)
                #print('-------------------------------------')
                

                ### Create Dataset Excel Sheet

#                wb = load_workbook(filename=excel_template)
                #ws = wb["GDP Metadaten"]
                #for dv in ws.data_validations.dataValidation:            
                #    print(dv.sqref)
                #    print(dv)
                #    print('-------')
                # Titel
                coords = 'B2'
                cell = ws[coords]
                cell.value = package_dict['title']
                # Beschreibung
                coords = 'B4'
                cell = ws[coords]
                cell.value = package_dict['notes']
                # Letzter Import
                coords = 'B30'
                cell = ws[coords]
                cell.value = package_dict['letzter_import']
                # Tags
                for key_ in package_dict.keys():
                    if 'tags.' in key_:
                        nr_ = int(key_.split('.')[1])
                        row_ = chr(65 + nr_ + 1)
                        coords = f'{row_}5'
                        cell = ws[coords]
                        cell.value = package_dict[key_]
                # Gruppen
                topic_categories = 0
                for key_ in package_dict.keys():
                    # Hauptkategorien
                    if 'groups.main' in key_:
                        coords = 'B29'
                        cell = ws[coords]
                        cell.value = package_dict[key_]
                        dv = add_validation_dropdown(ws, coords, '=Wertelisten!$A$2:$A$7')
                        ws.add_data_validation(dv)
                    # Themen    
                    if 'groups.topic' in key_:
                        nr_ = int(key_.split('.')[2])
                        topic_categories = topic_categories + 1
                        row_ = chr(65 + nr_ + 1)
                        coords = f'{row_}6'
                        cell = ws[coords]
                        cell.value = package_dict[key_]
                        dv = add_validation_dropdown(ws, coords, '=Wertelisten!$A$8:$A$23')
                        ws.add_data_validation(dv)                     
                # Add validation to empty fields in groups - topics:
                for l in range(topic_categories + 1, 11):
                    row_ = chr(65 + l)
                    coords = f'{row_}6'
                    cell = ws[coords]
                    dv = add_validation_dropdown(ws, coords, '=Wertelisten!$A$8:$A$23')
                    ws.add_data_validation(dv)
                # Organisation
                coords = 'B10'
                cell = ws[coords]
                cell.value = package_dict['organization.title']
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$B$2:$B$36')
                ws.add_data_validation(dv)
                # Privat
                #coords = 'B30'
                #cell = ws[coords]
                #if package_dict['private'] == False:
                #    cell.value = 'Nein'
                #elif package_dict['private'] == True:
                #    cell.value = 'Ja'
                #dv = add_validation_dropdown(ws, coords, '=Wertelisten!$M$2:$M$3')
                #ws.add_data_validation(dv)
                # Datenverantwortlicher
                coords = 'B11'
                cell = ws[coords]
                cell.value = package_dict['datenverantwortlicher']
                dv = add_validation_mandatory(ws, coords)
                ws.add_data_validation(dv)
                # Datenverantwortlicher Email
                coords = 'B12'
                cell = ws[coords]
                cell.value = package_dict['datenverantwortlicher_email']
                dv = add_validation_mandatory(ws, coords)
                ws.add_data_validation(dv)
                # Externe Datenurheber
                coords = 'B13'
                cell = ws['B13']
                cell.value = package_dict['ext_org']
                dv = add_validation_mandatory(ws, coords)
                ws.add_data_validation(dv)
                # Schema
                coords = 'B3'
                cell = ws[coords]
                cell.value = tp.schema_transpose_r[package_dict['schema']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$C$2:$C$21')
                ws.add_data_validation(dv)
                # Objekttyp
                coords = 'B7'
                cell = ws[coords]
                cell.value = tp.objekttyp_transpose_r[package_dict['objekttyp']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$D$2:$D$8')
                ws.add_data_validation(dv)
                # Update-Zyklus
                coords = 'B14'
                cell = ws[coords]
                cell.value = tp.update_zyklus_transpose_r[package_dict['update_zyklus']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$E$2:$E$8')
                ws.add_data_validation(dv)
                # Stadtarchiv - Archivwürdigkeit
                coords = 'B15'
                cell = ws[coords]
                cell.value = tp.archivwuerdigkeit_transpose_r[package_dict['stadtarchiv.archivwuerdigkeit']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$F$2:$F$4')
                ws.add_data_validation(dv)
                # Stadtarchiv - Begründung
                coords = 'B16'
                cell = ws[coords]
                cell.value = package_dict['stadtarchiv.begruendung']
                # Genauigkeit - Lagegenauigkeit
                coords = 'B8'
                cell = ws[coords]
                cell.value = tp.lagegenauigkeit_transpose_r[package_dict['genauigkeit.lagegenauigkeit']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$G$2:$G$11')
                ws.add_data_validation(dv)
                # Genauigkeit - Bemerkung
                coords = 'B9'
                cell = ws[coords]
                cell.value = package_dict['genauigkeit.bemerkung']
                # LHM-intern
                coords = 'B17'
                cell = ws[coords]
                cell.value = tp.lhm_intern_transpose_r[package_dict['lhm_intern']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$H$2:$H$4')
                ws.add_data_validation(dv)
                # MB_CHANGE FOR DOWNLOAD BUTTON ###
                # LHM-intern - GeoInfoWeb
                if type_ == 'geodatenpool':
                    coords = 'B19'
                    cell = ws[coords]
                    cell.value = tp.geoinfoweb_transpose_r[package_dict['geoinfoweb']]
                    dv = add_validation_dropdown(ws, coords, '=Wertelisten!$I$2:$I$6')
                    ws.add_data_validation(dv)
                ###############################
                # Internet - Veröffentlichungen LHM
                #coords = 'B21'
                #cell = ws[coords]
                #cell.value = tp.internet_praesentationen_transpose_r[package_dict['internet_praesentationen']]
                #dv = add_validation_dropdown(ws, coords, '=Wertelisten!$J$2:$J$4')
                #ws.add_data_validation(dv)
                # LHM-extern
                coords = 'B21'
                cell = ws[coords]
                cell.value = tp.lhm_extern_transpose_r[package_dict['lhm_extern']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$J$2:$J$4')
                ws.add_data_validation(dv)
                # LHM-extern - Nutzungsoptionen bei Einschränkungen
                line = '22'
                ind = 0
                cols_nu = ["B", "C", "D", "E"]
                if 'nutzungsoptionen_list' in package_dict.keys():
                    for col_nu in cols_nu:
                        coords = f'{col_nu}{line}'
                        cell = ws[coords]
                        if len(package_dict['nutzungsoptionen_list']) > ind:
                            option = package_dict['nutzungsoptionen_list'][ind]
                            cell.value = tp.nutzungsoptionen_transpose_r[option]
                        dv = add_validation_dropdown(ws, coords, '=Wertelisten!$K$2:$K$6')
                        ws.add_data_validation(dv)
                        ind = ind + 1
                else:
                    # Add validation for empty cells
                    for col_nu in cols_nu:
                        coords = f'{col_nu}{line}'
                        cell = ws[coords]
                        dv = add_validation_dropdown(ws, coords, '=Wertelisten!$K$2:$K$6')
                        ws.add_data_validation(dv)
                        ind = ind + 1

                # LHM-extern - Open Data
                coords = 'B23'
                cell = ws[coords]
                cell.value = tp.open_data_transpose_r[package_dict['open_data']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$L$2:$L$4')
                ws.add_data_validation(dv)

                # LHM-extern - High Value Datasets (HVD) 
                coords = 'B24'
                cell = ws[coords]
                cell.value = tp.hvd_transpose_r[package_dict['hvd']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$M$2:$M$4')
                ws.add_data_validation(dv)

                # LHM-extern - HVD-Kategorie
                coords = 'B25'
                cell = ws[coords]
                cell.value = tp.hvd_kategorie_transpose_r[package_dict['hvd_kategorie']]
                dv = add_validation_dropdown(ws, coords, '=Wertelisten!$N$2:$N$8')
                ws.add_data_validation(dv)

                # LHM-extern - Open Data / HVD - bestimmte Attribute
                coords = 'B26'
                cell = ws[coords]
                cell.value = package_dict['open_data_attribute']
                
                # Internet - Weitergabe im Auftrag
                #coords = 'B24'
                #cell = ws[coords]
                #cell.value = package_dict['datenabgabe_extern_mit_auftrag']
                #dv = add_validation_dropdown(ws, coords, '=Wertelisten!$M$2:$M$3')
                #ws.add_data_validation(dv)
                # Internet - Datenabgabe Externe
                #coords = 'B25'
                #cell = ws[coords]
                #cell.value = tp.datenabgabe_extern_transpose_r[package_dict['datenabgabe_extern']]
                #dv = add_validation_dropdown(ws, coords, '=Wertelisten!$L$2:$L$4')
                #ws.add_data_validation(dv)
                
                # Nutzungshinweise - Stichwort
                for key_ in package_dict.keys():
                    if 'nutzungshinweise.' in key_ and '.stichwort' in key_:
                        nr_ = int(key_.split('.')[1])
                        row_ = chr(65 + nr_ + 1)
                        coords = f'{row_}27'
                        cell = ws[coords]
                        cell.value = package_dict[key_]
                # Nutzungshinweise - Hinweise
                for key_ in package_dict.keys():
                    if 'nutzungshinweise.' in key_ and '.hinweise' in key_:
                        nr_ = int(key_.split('.')[1])
                        row_ = chr(65 + nr_ + 1)
                        coords = f'{row_}28'
                        cell = ws[coords]
                        cell.value = package_dict[key_]
                # MB_CHANGE FOR DOWNLOAD BUTTON ###
                # Add Sichtbarkeit, maybe remove row Geoinfoweb
                # Replace sheet name 'GDP_Metadaten' with 'Metadaten' (for all schemes) or 'Mobidam_Metadaten' (for Mobidam)
                # Replace PDF Header (Exportieren) with 'Metadaten' (for all schemes) or 'Mobidam_Metadaten' (for Mobidam)
                ###################################



                


                #####
                #print('#####################')
                #for dv in ws.data_validations.dataValidation:            
                #    print(dv.sqref)
                #    print(dv)
                #    print('-------')
                #print('#####################')

                # Save Excel
                name_ = package_dict['name']
                wb.save(f'{wdir}/excel/{name_}.xlsx')

            else:
                print('...')                       

# Get packages data

if __name__ == "__main__":

#    transfer_outdir = 'y'
#   
    config_ = config()
    api_key = config_[0]
    base_url = config_[1]
#
#    # a) via package list  
#    package_list = ckan_request('/api/3/action/package_list')
#
##    # Check data
#    private = ['waermeplanung_baublock', 'altlasten_flaechenumgriffe', 'altlasten', 'hausmaxpegel2017']
##    i = 0
##    for el in private:
##        i = i + 1
##        if i < 1000:
##            #print(el)
##            package_dict = ckan_request(f'/api/3/action/package_show?id={el}')
##            #print(package_dict['title'])
##            if ' | ' in package_dict['title']:
##                if not package_dict['title'].split(' | ')[1].lower() == el:
##                    print(el, 'a')
##                if not len(package_dict['title'].split(' | ')) == 2:
##                    print(el, 'b')
##            else:
##                print(el, 'c')
##    print(i)
#
#    #print(package_list)
#    packages = packages_to_files(package_list, 1000, 'gdp_prod_05', 'schema_template/template_v1.2.1.xlsx')
#    #packages = packages_to_files('richtwerte_2018', 10, 'gdp_dev_v1.2.1_02', 'schema_template/template_v1.2.1.xlsx')
#
#    # To Netzlaufwerk
#    if transfer_outdir == 'y':
#        print('Copy files to network drive:')
#        wdir = 'gdp_prod_05'
#        outdir = '/mnt/l/KOM/GDP_Metadaten'
#        #outdir = 'outdir'
#        for file in os.listdir(f'{wdir}/excel'):
#            print(file)
#            if os.path.isfile(f'{wdir}/excel' + '/' + file):
#                #shutil.copy(f'gdp_prod_v1.2.1/excel/{file}', f'/mnt/l/KOM/GDP_Metadaten/aktuell/{file}')
#                #shutil.copyfile(f'gdp_prod_v1.2.1/excel/{file}', f'aktuell/{file}')
#                shutil.copy2(f'{wdir}/excel/{file}', f'{outdir}/aktuell/{file}')
#                os.chmod(f'{outdir}/aktuell/{file}', 0o777)
#            elif os.path.isdir(f'{wdir}/excel' + '/' + file):
#                dir = file
#                shutil.copytree(f'{wdir}/excel/{dir}', f'{outdir}/aktuell/{dir}', dirs_exist_ok=True)
#                os.chmod(f'{outdir}/aktuell/{dir}', 0o777)
#                for file in os.listdir(f'{outdir}/aktuell/{dir}'):
#                    os.chmod(f'{outdir}/aktuell/{dir}/{file}', 0o777)


    # b) via package id
    #packages = packages_to_files('enp_erg_aggr_blo_poly', 1, 'gdp_test_v1.2.1_prod_selection', 'schema_template/template_v1.2.1.xlsx')
    #packages = packages_to_files('bauwerke_polygon_akt', 1, 'gdp_test_v1.2.1_prod_selection', 'schema_template/template_v1.2.1.xlsx')
    #packages = packages_to_files('richtwerte_2016', 1, 'gdp_test_v1.2.1_prod_selection', 'schema_template/template_v1.2.1.xlsx')

    packages = packages_to_files('rad_vepr_zone30_poly', 1, 'gdp_upgrade_dataset_3', 'schema_template/template_v1.2.1_upgrade.xlsx')
    



